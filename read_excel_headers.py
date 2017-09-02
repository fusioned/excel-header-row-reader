#!/usr/bin/python3
"""
Excel workbook header line bulk reader
======================================

Reads and outputs the header row for the given Excel files
Optionally append the first content row after the header for additional inspection

"""

import sys
import argparse
import csv
from openpyxl import load_workbook

class CsvWriterRender:
    """Output to a CSV format using the csv module """

    def __init__(self, output=sys.stdout, num_cols_pad=20, append_first_body_row=False):
        """
        num_cols_pad indicate the number of columns minimum before the body data is added on
        e.g. when num_cols_pad = 3 and append_first_body_row = True
            ,,,extra,body_content
            a,b,,extra,content
            filename,,,additional

        """
        self.num_cols = num_cols_pad
        self.append_body = append_first_body_row
        self.row = []
        self.output = csv.writer(output)

    def start(self):
        pass

    def filename(self, filename):
        self._cell(filename)

    def sheet_name(self, name):
        self._cell(name)

    def header_cell(self, value, cell, index):
        self._cell(value if value else '')

    def header_row(self, row):
        pass

    def body_row(self, row):
        # pad this out
        if self.append_body:
            self.row += [''] * (self.num_cols - len(self.row))

    def body_cell(self, value, cell, index):
        if self.append_body:
            self._cell(value if value else '')

    def done(self):
        self._newline()

    def _cell(self, value):
        self.row += [value]

    def _newline(self):
        if self.row:
            self.output.writerow(self.row)
        self.row = []


class ClassicRender:
    """Output data indented"""

    def __init__(self, append_first_body_row=False):
        """
        Option to append the first content row
        """
        self.append_body = append_first_body_row

    def start(self):
        pass

    def filename(self, filename):
        print('[' + filename + ']')

    def sheet_name(self, name):
        print(name + ':')

    def header_cell(self, value, cell, index):
        if value:
            print('  ' + value)

    def header_row(self, row):
        pass

    def body_row(self, row):
        if (self.append_body):
            for cell in row:
                print('  >  ' + (cell.value if cell.value else ''))

    def body_cell(self, value, cell, index):
        pass

    def done(self):
        print('')


def main():
    args = parse_args()

    if (args.format == 'classic'):
        renderer = ClassicRender(append_first_body_row=args.body)
    else:
        renderer = CsvWriterRender(append_first_body_row=args.body, 
            output=args.output,
            num_cols_pad=args.pad_left or 20)

    for filename in args.filenames:
        process(filename, renderer)

def parse_args():
    parser = argparse.ArgumentParser(description='Parse Excel workbooks for the headers')
    parser.add_argument('filenames', metavar='workbooks', help='Workbook files', nargs='+')
    parser.add_argument('--include-body', '-b', dest='body', action='store_true',
                       help='Append the next row of content data to the header output')
    parser.add_argument('--pad-body', '-p', dest='pad_left', type=int,
                       help='The number of cells to allocate for the headers (when using the CSV output format)')
    parser.add_argument('--format', '-f', dest='format',
                       help='Output format - this can be: csv, classic')
    parser.add_argument('--output', nargs='?', type=argparse.FileType('w'), 
                        default=sys.stdout)

    args = parser.parse_args()

    return(args)


def process(filename, renderer):
    renderer.filename(filename)
    wb = load_workbook(filename)
    sheets = wb.get_sheet_names()

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        renderer.sheet_name(sheet_name)

        header = sheet[1]
        renderer.header_row(header)
        for index, cell in enumerate(header):
            renderer.header_cell(cell.value, cell, index)

        row = sheet[2]
        renderer.body_row(row)
        for index, cell in enumerate(row):
            renderer.body_cell(cell.value, cell, index)

    renderer.done()


if __name__ == '__main__':
    main()
